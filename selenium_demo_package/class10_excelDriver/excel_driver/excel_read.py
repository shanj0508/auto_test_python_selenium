# -*- coding: utf-8 -*-
'''
    Excel数据驱动效果实现。
    实现目的：基于excel中的内容，来调用关键字类，实现自动化测试的执行效果
            相当于，excel文件就是一个测试用例，底层代码就是关键字驱动类以及excel驱动类
'''

import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from selenium_demo_package.class10_excelDriver.excel_driver.web_keywordDriven import WebKey

# 读取excel文件，创建workbook实例化对象


excel = openpyxl.load_workbook('../excel_data/data.xlsx')

# 设置单元格样式
# 定义边框样式 border
border_type_pass = Side(border_style=None, color='C6EFCE')
border_pass = Border(left=border_type_pass, right=border_type_pass, top=border_type_pass, bottom=border_type_pass)

border_type_failed = Side(border_style=None, color='FFC7CE')
border_failed = Border(left=border_type_failed, right=border_type_failed, top=border_type_failed,
                       bottom=border_type_failed)

# 定义字体样式 font
font_pass = Font(size=11, bold=True, name='宋体', color="006100")  # 字体大小，加粗，字体名称，字体颜色
font_failed = Font(size=11, bold=True, name='宋体', color="9C0006")  # 字体大小，加粗，字体名称，字体颜色

# 定义填充样式
fill_pass = PatternFill(patternType="solid", start_color="C6EFCE")  # 纯色填充
fill_failed = PatternFill(patternType="solid", start_color="FFC7CE")  # 纯色填充

# 定义对齐方式
align = Alignment(horizontal='center', vertical='center', wrap_text=True)  # 居中对齐

# 获取所有的sheet页
sheets = excel.sheetnames
# 遍历所有sheet页
for sheet in sheets:
    sheet_temp = excel[sheet]
    # 如果sheet中包含特定字段，直接continue
    # 遍历sheet页中所有单元格
    for values in sheet_temp.values:
        # 读取用例的执行部分内容，表头的一些描述内容不需要读取，即从编号为1开始读取
        # print(values[1])
        # print(type(values[1]))
        if type(values[0]) is int:
            # print(values)

            '''用例结构：
                    1.用例编号	
                    2.执行操作：调用的关键字，结合关键字驱动类应用
                    3.定位方法：关键字传入的参数，结合关键字驱动类应用
                    4.定位路径：关键字传入的参数，结合关键字驱动类应用
                    5.输入文本：关键字传入的参数，结合关键字驱动类应用	
                    6.操作描述：可以输出到日志中
            '''
            # 提取本行的测试数据
            data = {}
            data['name'] = values[2]  # 定位方法
            data['value'] = values[3]  # 定位路径
            data['text'] = values[4]  # 输入文本
            # print(data)
            '''data输出结果
                {'text': u'Chrome', 'name': None, 'value': None}
                {'text': u'http://39.98.138.157/shopxo/', 'name': None, 'value': None}
                {'text': None, 'name': u'link text', 'value': u'\u767b\u5f55'}
                {'text': None, 'name': u'xpath', 'value': u'//*[text()="\u8fd8\u6ca1\u6709\u5e10\u53f7\uff1f"]'}
                {'text': u'shanjing', 'name': u'name', 'value': u'accounts'}
                {'text': u'shanjing@.000', 'name': u'name', 'value': u'pwd'}
                {'text': None, 'name': u'xpath', 'value': u'//button[text()="\u767b\u5f55"]'}
                {'text': None, 'name': u'xpath', 'value': u'//*[text()="\u767b\u5f55\u6210\u529f"]'}
                {'text': u'\u9000\u51fa', 'name': u'link text', 'value': u'\u9000\u51fa'}
                {'text': 5L, 'name': None, 'value': None}
                {'text': None, 'name': None, 'value': None}

            '''
            # 优化测试数据内容，将所有值为None的数据全部清除出data中
            for key in list(data.keys()):
                # print(key)
                if data[key] is None:
                    del data[key]
            print(values[5])
            print(data)
            '''data输出结果
                {'text': u'Chrome'}
                {'text': u'http://39.98.138.157/shopxo/'}
                {'name': u'link text', 'value': u'\u767b\u5f55'}
                {'name': u'xpath', 'value': u'//*[text()="\u8fd8\u6ca1\u6709\u5e10\u53f7\uff1f"]'}
                {'text': u'shanjing', 'name': u'name', 'value': u'accounts'}
                {'text': u'shanjing@.000', 'name': u'name', 'value': u'pwd'}
                {'name': u'xpath', 'value': u'//button[text()="\u767b\u5f55"]'}
                {'name': u'xpath', 'value': u'//*[text()="\u767b\u5f55\u6210\u529f"]'}
                {'text': u'\u9000\u51fa', 'name': u'link text', 'value': u'\u9000\u51fa'}
                {'text': 5L}
                {}
            '''
            # 调用对应的关键字来执行操作行为
            '''
            执行操作分为三种场景：
                1.open_browser：创建实例化类对象WebKey
                2.assert_text：断言，返回True或False,将结果写入excel的实际结果单元格中
                3.其他常规操作：访问、点击、输入等
            每种不同的场景，需要不同的处理。
        '''

            if values[1] == 'open_browser':
                # 实例化对象
                wk = WebKey(values[4])
                # wk = WebKey(**data)
            # 断言可能有多种，例如文本断言、属性断言等，对应多个断言方法。
            # 因此这里判断：values[1]的值中包含'assert_'即表示是个断言
            elif 'assert_' in values[1]:
                # 用变量接收断言的返回值：True或False
                status = getattr(wk, values[1])(**data)
                # 基于status写入测试结果
                if status:
                    # 写入Pass
                    sheet_temp.cell(row=values[0] + 2, column=7).value = 'Pass'
                    sheet_temp.cell(row=values[0] + 2, column=7).border = border_pass
                    sheet_temp.cell(row=values[0] + 2, column=7).font = font_pass
                    sheet_temp.cell(row=values[0] + 2, column=7).fill = fill_pass
                    sheet_temp.cell(row=values[0] + 2, column=7).alignment = align

                else:
                    # 写入Failed
                    sheet_temp.cell(row=values[0] + 2, column=7).value = 'Failed'
                    sheet_temp.cell(row=values[0] + 2, column=7).border = border_failed
                    sheet_temp.cell(row=values[0] + 2, column=7).font = font_failed
                    sheet_temp.cell(row=values[0] + 2, column=7).fill = fill_failed
                    sheet_temp.cell(row=values[0] + 2, column=7).alignment = align
            else:
                # 利用反射
                getattr(wk, values[1])(**data)
                # 上面这句等价于：wk.attr(**data)
                # values[1]是excel中的第一列，即：操作步骤
                '''
                    正常逻辑是：
                    if values[1]=='input':
                        wk.input(**data)
                    elif values[1]=='click':
                        wk.click(**data)
                        # 拆包:上面这行拆解为下面这行
                        # wk.click(name=xxx,value=xxx)
                    按照这种逻辑下，我们需要把所有的关键字都做成if判断，不然会出错，这样会导致代码量非常巨大
                    因此python提供了这种反射机制，来简化代码。
                '''

# 执行excel保存
excel.save('../excel_data/data.xlsx')
