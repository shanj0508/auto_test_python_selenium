# -*- coding: utf-8 -*-
import openpyxl

# excel数据读取:如果在操作时，方法无法点出来，就手动填写

# 1.读取excel文件
excel = openpyxl.load_workbook('../excel_data/data.xlsx')
# 2.读取excel中指定的sheet页
sheet = excel['Sheet1']
# 3.读取指定的单元格
# cell = sheet['A1']
# print(cell)  # <Cell u'Sheet1'.A1>
# print(cell.value)  # 自动化1


# 读取整个sheet页的内容：自动获取最大的行数和列数，没有内容的单元格用None表示
# values = sheet.values
# for v in values:
#     print(v) # v是元组的格式，可以直接通过下标获取指定单元格的内容  print(v[0])
# 打印结果
# (u'test1', None, None, None)
# (None, u'test2', None, None)
# (None, None, u'test3', None)
# (None, None, None, u'test4')

# 获取所有的sheet页中全部单元格内容
# 1.获取excel中所有的sheet页名称，返回list格式的内容
# sheets = excel.sheetnames
# print(sheets)
# # 2.通过循环读取sheets中的内容，以key的形式赋值给excel
# for sheet in sheets:
#     print(sheet)
#     print('*' * 20)
#     # 3.每一个sheet中的单元格内容输出
#     for value in excel[sheet].values:
#         print(value)

'''打印结果：
[u'Sheet1', u'Sheet2', u'Sheet3']
Sheet1
********************
(u'test1', None, None, None)
(None, u'test2', None, None)
(None, None, u'test3', None)
(None, None, None, u'test4')
Sheet2
********************
(u'sheet2',)
Sheet3
********************
(u'sheet3',)

'''

# excel数据写入:1.要先关闭excel文件 2.写完后需要执行保存操作
sheet.cell(row=5, column=5).value = 'test5'
excel.save('../excel_data/data.xlsx')
